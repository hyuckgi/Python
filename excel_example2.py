# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()

dest_filename = 'empty_book.xls'

ws = wb.active

ws.title = u'샘플'
ws['A1'] = '과일종류'
ws['A2'] = '메론'
ws['A3'] = '딸기'
ws['A4'] = '바나나'
ws['A5'] = '호박'
ws['A5'] = '과일종류'
ws['A6'] = '집계'
ws['B1'] = '재고'
ws['B2'] = 10
ws['B3'] = 2
ws['B4'] = 22
ws['B5'] = 10
ws['B6'] = 'SUM(B2:B5)'


wb.save(filename = dest_filename)
